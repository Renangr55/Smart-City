from django.shortcuts import render, HttpResponse
from .models import (Sensores,
                     Ambientes,
                     Historico)
from .serializer import (SensoresSerializer,
                        AmbientesSerializer,
                        HistoricoSerializer
)

from rest_framework.generics import ListCreateAPIView,RetrieveUpdateDestroyAPIView
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
import pandas as pd
import openpyxl

class PaginationConfig(PageNumberPagination):
    page_size = 3
    page_size_query_params = 'page_size'
    max_page_size = 10

# Create your views here.
class SensoresListCreateAPIView(ListCreateAPIView):
    queryset = Sensores.objects.all()
    serializer_class = SensoresSerializer
    pagination_class = PaginationConfig
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        sensor = self.request.query_params.get('sensor')
        if sensor:
            queryset = queryset.filter(nome__icontains=sensor)
        return queryset
    

class SensoresRetriveUpdateDestroyAPIView(RetrieveUpdateDestroyAPIView):
    queryset = Sensores.objects.all()
    serializer_class = SensoresSerializer
    pagination_class = PaginationConfig
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)
    
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(SensoresSerializer,data=request.data,partial=partial) 
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer) 

        if getattr(instance,'_prefetched_objects_cache', None):
            instance._prefetched_objects_cache = {}
        
        return Response(serializer.data)
    
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({'Mensagem' : 'Sensor Deletado'}, status=status.HTTP_204_NO_CONTENT)

class AmbientesListCreateAPIView(ListCreateAPIView):
    queryset = Ambientes.objects.all()
    serializer_class = AmbientesSerializer
    pagination_class = PaginationConfig
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        sig = self.request.query_params.get('sig')
        if sig:
            queryset = queryset.filter(nome__icontains=sig)
        return queryset
    

class AmbientesRetrieveUpdateDestroyAPIView(RetrieveUpdateDestroyAPIView):
    queryset = Ambientes.objects.all()
    serializer_class = AmbientesSerializer
    pagination_class = PaginationConfig
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response({'Mensagem' : 'A Classe: Ambiente foi listado'},serializer.data)
    
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(Ambientes,data=request.data,partial=partial) 
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer) 

        if getattr(instance,'_prefetched_objects_cache', None):
            instance._prefetched_objects_cache = {}
    
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({'Mensagem' : 'Ambiente deletado'}, status=status.HTTP_204_NO_CONTENT)


class HistoricoListCreateAPIView(ListCreateAPIView):
    queryset = Historico.objects.all()
    serializer_class = HistoricoSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = PaginationConfig
        

class HistoricoRetriveUpdateDestroyAPIView(RetrieveUpdateDestroyAPIView):
    queryset = Historico.objects.all()
    serializer_class = HistoricoSerializer
    pagination_class = PaginationConfig
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response({'Mensagem' : 'A Classe: Histórico foi listado' },serializer.data)
    
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(SensoresSerializer,data=request.data,partial=partial) 
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer) 

        if getattr(instance,'_prefetched_objects_cache', None):
            instance._prefetched_objects_cache = {}

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({"Mensagem" : "Histórico deletado com sucesso"})
    
def ImportarDadosExcel(request):
    wb_contador = openpyxl.load_workbook("../Dados_Integrador/contador.xlsx")
    wb_luminosidade = openpyxl.load_workbook("../Dados_Integrador/luminosidade.xlsx")
    wb_temperatura = openpyxl.load_workbook("../Dados_Integrador/temperatura.xlsx")
    wb_umidade = openpyxl.load_workbook("../Dados_Integrador/umidade.xlsx")

    planilha_contador = wb_contador.active
    planilha_luminosidade = wb_luminosidade.active
    planilha_temperatura = wb_temperatura.active
    planilha_umidade = wb_umidade.active
    

    
    InfoSensores = []

    Sensores.objects.all().delete()
    

    

    planilhaSensores = [
        planilha_contador,
        planilha_luminosidade, 
        planilha_temperatura,
        planilha_umidade
    ]

    for planilha in planilhaSensores:
        for row in planilha.iter_rows(min_row=2, values_only=True):
            print(row)
            if len(row) >= 5:
                sensores = Sensores(
                    sensor=row[0],
                    mac_address=row[1],
                    unidade_med=row[2],
                    latitude=row[3],
                    longitude=row[4],
                    status=row[5],
                )
                InfoSensores.append(sensores)
                
            else:
                return HttpResponse ("Colunas insulficiente")
            
    Sensores.objects.abulk_create(InfoSensores)

    return HttpResponse ("IMPORTANDO DADOS", InfoSensores)


def ImportarDadosAmbiente(request):
        wb_ambiente = openpyxl.load_workbook("../Dados_Integrador/Ambientes.xlsx")
        planilha_ambiente = wb_ambiente.active

        InfoAmbiente = []
        Ambientes.objects.all().delete()

        for row in planilha_ambiente.iter_rows(min_row=2, values_only=True):
            ambientes = Ambientes(
                sig=row[0],
                descricao=row[1],
                ni=row[2],
                responsavel=row[3]
            ) 
            InfoAmbiente.append(ambientes)
        
        Ambientes.objects.abulk_create(InfoAmbiente)
        return HttpResponse ("IMPORTANDO DADOS", InfoAmbiente)










    

    





        





    
        



        
    